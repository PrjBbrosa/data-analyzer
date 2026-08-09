from __future__ import annotations

from pathlib import Path
import importlib
import json
import os
import subprocess
import sys

import pytest

import mf4_analyzer.batch_output as batch_output
from mf4_analyzer.batch_output import (
    OutputPublishRace,
    OutputRollbackIncomplete,
    atomic_write,
    atomic_write_set,
    build_task_output_identity,
    build_frf_task_output_identity,
    choose_output_paths,
    inspect_output_reservation,
    release_output_reservation,
    reserve_output_paths,
    unicode_slug,
)


class _Source:
    def __init__(self, filepath: Path, label_suffix: str = ""):
        self.filepath = filepath
        self.label_suffix = label_suffix
        self.source_metadata = {}


def _group_members():
    return (
        ("/runs/a.mf4", "group-a", "speed"),
        ("/runs/b.mf4", "group-b", "speed"),
    )


def test_xlsx_writer_splits_only_at_excel_row_limit(tmp_path, monkeypatch):
    import pandas as pd
    from openpyxl import load_workbook

    import mf4_analyzer.batch as batch_module

    monkeypatch.setattr(batch_module, "_XLSX_MAX_DATA_ROWS", 2)
    target = tmp_path / "result.xlsx"
    batch_module.BatchRunner._write_dataframe(
        pd.DataFrame({"sample": [1, 2, 3]}), target,
    )

    book = load_workbook(target, read_only=True)
    assert book.sheetnames == ["数据1", "数据2"]
    assert list(book["数据1"].values) == [("sample",), (1,), (2,)]
    assert list(book["数据2"].values) == [("sample",), (3,)]


def test_group_identity_changes_when_one_member_source_changes():
    first = batch_output.build_group_output_identity(
        _group_members(), method="time",
        params={"render_group_by": "channel"}, group_by="channel",
    )
    changed = batch_output.build_group_output_identity(
        (
            ("/archive/a.mf4", "group-a", "speed"),
            ("/runs/b.mf4", "group-b", "speed"),
        ),
        method="time",
        params={"render_group_by": "channel"},
        group_by="channel",
    )

    assert first.group_id != changed.group_id
    assert first.stem != changed.stem


def test_group_identity_is_member_order_independent():
    first = batch_output.build_group_output_identity(
        _group_members(), method="time",
        params={"render_group_by": "channel"}, group_by="channel",
    )
    reversed_members = batch_output.build_group_output_identity(
        tuple(reversed(_group_members())), method="time",
        params={"render_group_by": "channel"}, group_by="channel",
    )

    assert first == reversed_members
    assert first.members == tuple(sorted(_group_members()))


@pytest.mark.parametrize(
    "changed_members",
    (
        (
            ("/runs/a.mf4", "other-group", "speed"),
            ("/runs/b.mf4", "group-b", "speed"),
        ),
        (
            ("/runs/a.mf4", "group-a", "rpm"),
            ("/runs/b.mf4", "group-b", "speed"),
        ),
    ),
)
def test_group_identity_uses_complete_member_identity(changed_members):
    base = batch_output.build_group_output_identity(
        _group_members(), method="time",
        params={"render_group_by": "channel"}, group_by="channel",
    )
    changed = batch_output.build_group_output_identity(
        changed_members, method="time",
        params={"render_group_by": "channel"}, group_by="channel",
    )

    assert base.group_id != changed.group_id


def test_group_render_tasks_orders_source_and_channel_groups_deterministically():
    grouping = importlib.import_module("mf4_analyzer.batch_grouping")
    tasks = (
        grouping.RenderTask(
            "source-b", "rpm",
            batch_output.TaskOutputIdentity(
                "b-rpm", "/runs/b.mf4", "group-b", "rpm", "b-rpm-stem",
            ),
        ),
        grouping.RenderTask(
            "source-a", "speed",
            batch_output.TaskOutputIdentity(
                "a-speed", "/runs/a.mf4", "group-a", "speed", "a-speed-stem",
            ),
        ),
        grouping.RenderTask(
            "source-b", "speed",
            batch_output.TaskOutputIdentity(
                "b-speed", "/runs/b.mf4", "group-b", "speed", "b-speed-stem",
            ),
        ),
        grouping.RenderTask(
            "source-a", "rpm",
            batch_output.TaskOutputIdentity(
                "a-rpm", "/runs/a.mf4", "group-a", "rpm", "a-rpm-stem",
            ),
        ),
    )

    source_groups = grouping.group_render_tasks(
        tuple(reversed(tasks)),
        {"render_group_by": "source", "render_layout": "subplot"},
    )
    channel_groups = grouping.group_render_tasks(
        tasks, {"render_group_by": "channel"},
    )
    singleton_groups = grouping.group_render_tasks(tasks, {})

    assert [
        [member.identity.task_id for member in group.members]
        for group in source_groups
    ] == [["a-rpm", "a-speed"], ["b-rpm", "b-speed"]]
    assert [group.layout for group in source_groups] == ["subplot", "subplot"]
    assert [
        [member.identity.task_id for member in group.members]
        for group in channel_groups
    ] == [["a-rpm", "b-rpm"], ["a-speed", "b-speed"]]
    assert [group.identity.stem for group in singleton_groups] == [
        "a-rpm-stem", "a-speed-stem", "b-rpm-stem", "b-speed-stem",
    ]


def test_default_time_render_params_preserve_task_identity(tmp_path):
    source = _Source(tmp_path / "legacy.mf4", "default")
    legacy = build_task_output_identity(
        source, file_id=1, channel="speed", method="time", params={},
    )
    explicit_defaults = build_task_output_identity(
        source,
        file_id=1,
        channel="speed",
        method="time",
        params={
            "render_group_by": "none",
            "render_layout": "overlay",
            "x_source": "time",
            "x_channel": "",
            "x_origin": "zero",
        },
    )

    assert explicit_defaults == legacy


def test_unicode_slug_keeps_cjk_and_removes_only_path_unsafe_characters():
    assert unicode_slug("振动/通道:左侧") == "振动_通道_左侧"


def test_frf_task_identity_is_directional_and_pair_aware(tmp_path):
    source = _Source(tmp_path / "试验:01.mf4", "default")
    outputs = {
        "export_data": True,
        "export_image": True,
        "data_format": "csv",
        "image_format": "png",
    }
    forward = build_frf_task_output_identity(
        source,
        file_id="logical-a",
        input_channel="命令/输入",
        output_channel="响应:输出",
        params={"estimator": "h1", "frequency_scale": "log"},
        outputs=outputs,
    )
    repeated = build_frf_task_output_identity(
        source,
        file_id="ignored-when-path-exists",
        input_channel="命令/输入",
        output_channel="响应:输出",
        params={"estimator": "h1", "frequency_scale": "log"},
        outputs=outputs,
    )
    reverse = build_frf_task_output_identity(
        source,
        file_id="logical-a",
        input_channel="响应:输出",
        output_channel="命令/输入",
        params={"estimator": "h1", "frequency_scale": "log"},
        outputs=outputs,
    )

    assert forward == repeated
    assert forward.task_id != reverse.task_id
    assert forward.input_channel_identity == "命令/输入"
    assert forward.output_channel_identity == "响应:输出"
    assert forward.channel_identity == "响应:输出 / 命令/输入"
    assert forward.stem.startswith("试验_01__响应_输出-over-命令_输入__frf__")


def test_frf_coordinated_artifact_identity_includes_render_and_output_bytes(tmp_path):
    source = _Source(tmp_path / "run.mf4")
    base_outputs = {
        "export_data": True,
        "export_image": True,
        "data_format": "csv",
        "image_format": "png",
        "image_width": 1920,
    }

    def identity(params, outputs=base_outputs):
        return build_frf_task_output_identity(
            source,
            file_id=1,
            input_channel="cmd",
            output_channel="actual",
            params=params,
            outputs=outputs,
        )

    base = identity({"estimator": "h1", "coherence_threshold": 0.8})
    render_changed = identity({"estimator": "h1", "coherence_threshold": 0.5})
    output_changed = identity(
        {"estimator": "h1", "coherence_threshold": 0.8},
        {**base_outputs, "image_width": 2560},
    )

    assert base.task_id != render_changed.task_id
    assert base.task_id != output_changed.task_id


@pytest.mark.parametrize(
    ("input_channel", "output_channel"),
    (("", "out"), ("in", ""), ("same", "same")),
)
def test_frf_task_identity_rejects_missing_or_self_pair(
    tmp_path, input_channel, output_channel,
):
    with pytest.raises(ValueError):
        build_frf_task_output_identity(
            _Source(tmp_path / "source.mf4"),
            file_id=1,
            input_channel=input_channel,
            output_channel=output_channel,
            params={"nfft_mode": "auto"},
        )


def test_task_output_identity_is_stable_and_separates_source_and_group(tmp_path):
    first = build_task_output_identity(
        _Source(tmp_path / "a" / "同名.hdf", "1x"),
        file_id=1,
        channel="振动",
        method="fft",
        params={"nfft": 1024},
    )
    repeated = build_task_output_identity(
        _Source(tmp_path / "a" / "同名.hdf", "1x"),
        file_id=99,
        channel="振动",
        method="fft",
        params={"nfft": 1024},
    )
    other_source = build_task_output_identity(
        _Source(tmp_path / "b" / "同名.hdf", "1x"),
        file_id=1,
        channel="振动",
        method="fft",
        params={"nfft": 1024},
    )
    other_group = build_task_output_identity(
        _Source(tmp_path / "a" / "同名.hdf", "2x"),
        file_id=1,
        channel="振动",
        method="fft",
        params={"nfft": 1024},
    )

    assert first == repeated
    assert first.task_id != other_source.task_id
    assert first.task_id != other_group.task_id
    assert "同名__1x__振动__fft__" in first.stem


@pytest.mark.parametrize("label", ("", "   ", "default", " default "))
def test_task_stem_omits_the_uninformative_default_group(tmp_path, label):
    identity = build_task_output_identity(
        _Source(tmp_path / "repro.mf4", label),
        file_id=1,
        channel="sig",
        method="time",
        params={},
    )

    assert identity.group_identity == "default"
    assert "default" not in identity.stem
    assert identity.stem == f"repro__sig__time__{identity.task_id[:8]}"


def test_task_stem_keeps_a_real_group_label_after_the_source(tmp_path):
    identity = build_task_output_identity(
        _Source(tmp_path / "repro.mf4", "cycle-A"),
        file_id=1,
        channel="sig",
        method="time",
        params={},
    )

    assert identity.group_identity == "cycle-A"
    assert identity.stem == f"repro__cycle-A__sig__time__{identity.task_id[:8]}"


def test_task_stem_drops_a_blank_group_that_identity_still_carries(tmp_path):
    """Whitespace-only metadata stays in the hash but must not reach the name."""

    source = _Source(tmp_path / "repro.mf4")
    source.source_metadata = {"group": "   "}
    identity = build_task_output_identity(
        source, file_id=1, channel="sig", method="time", params={},
    )

    assert identity.group_identity == "   "
    assert identity.stem == f"repro__sig__time__{identity.task_id[:8]}"


def test_group_stem_drops_the_group_by_literal_and_the_default_group():
    by_source = batch_output.build_group_output_identity(
        (
            ("/runs/repro.mf4", "default", "sig"),
            ("/runs/repro.mf4", "default", "aux"),
        ),
        method="time",
        params={"render_group_by": "source"},
        group_by="source",
    )
    labelled_source = batch_output.build_group_output_identity(
        (
            ("/runs/repro.mf4", "cycle-A", "sig"),
            ("/runs/repro.mf4", "cycle-A", "aux"),
        ),
        method="time",
        params={"render_group_by": "source"},
        group_by="source",
    )
    by_channel = batch_output.build_group_output_identity(
        (
            ("/runs/a.mf4", "default", "aux"),
            ("/runs/b.mf4", "default", "aux"),
        ),
        method="time",
        params={"render_group_by": "channel"},
        group_by="channel",
    )

    assert by_source.stem == f"repro__time__{by_source.group_id[:8]}"
    assert "default" not in by_source.stem
    assert "source" not in by_source.stem.split("__")
    assert labelled_source.stem == (
        f"repro__cycle-A__time__{labelled_source.group_id[:8]}"
    )
    assert "source" not in labelled_source.stem.split("__")
    assert by_channel.stem == f"aux__time__{by_channel.group_id[:8]}"
    assert "channel" not in by_channel.stem.split("__")


def test_readable_stem_shortening_did_not_move_the_hashed_identities():
    """Filenames are cosmetic; these hashes drive manifests, resume and reuse.

    The golden values were captured from the pre-shortening implementation, so
    a diff here means an identity break, not merely a renamed artifact.
    """

    task = build_task_output_identity(
        _Source(None), file_id=7, channel="spd", method="fft",
        params={"fs": 1024.0},
    )
    group_source = batch_output.build_group_output_identity(
        (
            ("/runs/repro.mf4", "default", "sig"),
            ("/runs/repro.mf4", "default", "aux"),
        ),
        method="time",
        params={"render_group_by": "source"},
        group_by="source",
    )
    group_channel = batch_output.build_group_output_identity(
        (
            ("/runs/a.mf4", "default", "aux"),
            ("/runs/b.mf4", "default", "aux"),
        ),
        method="time",
        params={"render_group_by": "channel"},
        group_by="channel",
    )

    assert task.task_id == (
        "a57665c65252c07263ead552bd2e73614e43992d2614382c2c2439f5f6f3f7d5"
    )
    assert group_source.group_id == (
        "7def54cefc7b4235eda0c6e4da523bce65dfeb001d0f80101a0e3b800710a704"
    )
    assert group_channel.group_id == (
        "96e0d4bf0b3e28e7461cf7bbb29db0fa53bef40d1a3799815ed2543c2ff25012"
    )


def test_choose_output_paths_auto_numbers_without_overwriting(tmp_path):
    (tmp_path / "result.csv").write_text("old", encoding="utf-8")

    paths = choose_output_paths(tmp_path, "result", ("csv", "png"))

    assert paths["csv"].name == "result__2.csv"
    assert paths["png"].name == "result__2.png"
    assert (tmp_path / "result.csv").read_text(encoding="utf-8") == "old"


def test_atomic_write_replaces_only_after_success_and_cleans_own_temp(tmp_path):
    target = tmp_path / "result.csv"

    atomic_write(target, lambda temp: temp.write_text("complete", encoding="utf-8"))

    assert target.read_text(encoding="utf-8") == "complete"
    assert not list(tmp_path.glob(".result.*.csv"))


def test_atomic_write_failure_leaves_no_final_or_temp_file(tmp_path):
    target = tmp_path / "result.csv"

    def fail(temp):
        temp.write_text("partial", encoding="utf-8")
        raise RuntimeError("simulated write failure")

    with pytest.raises(RuntimeError, match="simulated"):
        atomic_write(target, fail)

    assert not target.exists()
    assert not list(tmp_path.glob(".result.*.csv"))


def test_reserve_output_paths_coordinates_auto_numbered_artifact_set(tmp_path):
    (tmp_path / "result.csv").write_text("old", encoding="utf-8")

    reservation = reserve_output_paths(
        tmp_path, "result", ("csv", "png"),
        conflict_policy="auto_number",
    )
    try:
        assert reservation.status == "reserved"
        assert reservation.paths["csv"].name == "result__2.csv"
        assert reservation.paths["png"].name == "result__2.png"
    finally:
        reservation.release()


def test_active_reservation_forces_competing_auto_number_to_next_suffix(tmp_path):
    first = reserve_output_paths(tmp_path, "result", ("csv", "png"))
    second = reserve_output_paths(tmp_path, "result", ("csv", "png"))
    try:
        assert first.paths["csv"].name == "result.csv"
        assert second.paths["csv"].name == "result__2.csv"
        assert second.paths["png"].name == "result__2.png"
    finally:
        second.release()
        first.release()


def test_conflict_error_skip_and_overwrite_have_frozen_task_level_semantics(tmp_path):
    (tmp_path / "result.csv").write_text("old", encoding="utf-8")

    with pytest.raises(FileExistsError, match="already exists"):
        reserve_output_paths(
            tmp_path, "result", ("csv", "png"), conflict_policy="error",
        )

    skipped = reserve_output_paths(
        tmp_path, "result", ("csv", "png"), conflict_policy="skip",
    )
    assert skipped.status == "skipped"
    assert "manifest" in skipped.warning
    assert skipped.paths["csv"].name == "result.csv"

    overwrite = reserve_output_paths(
        tmp_path, "result", ("csv", "png"), conflict_policy="overwrite",
    )
    try:
        assert overwrite.status == "reserved"
        assert overwrite.paths["csv"].name == "result.csv"
        assert overwrite.paths["png"].name == "result.png"
        assert (tmp_path / "result.csv").read_text(encoding="utf-8") == "old"
    finally:
        overwrite.release()


def test_atomic_write_set_stages_every_writer_before_overwrite_publication(tmp_path):
    csv_path = tmp_path / "result.csv"
    png_path = tmp_path / "result.png"
    csv_path.write_text("old-csv", encoding="utf-8")
    png_path.write_bytes(b"old-png")
    reservation = reserve_output_paths(
        tmp_path, "result", ("csv", "png"), conflict_policy="overwrite",
    )

    def fail_image(path):
        path.write_bytes(b"partial-new-image")
        raise RuntimeError("render failed")

    with pytest.raises(RuntimeError, match="render failed"):
        atomic_write_set(
            reservation,
            {
                "csv": lambda path: path.write_text("new-csv", encoding="utf-8"),
                "png": fail_image,
            },
        )

    assert csv_path.read_text(encoding="utf-8") == "old-csv"
    assert png_path.read_bytes() == b"old-png"
    assert not list(tmp_path.glob(".*.batch-stage.*"))


def test_atomic_write_set_publishes_coordinated_suffix_and_releases_lock(tmp_path):
    (tmp_path / "result.csv").write_text("old", encoding="utf-8")
    reservation = reserve_output_paths(
        tmp_path, "result", ("csv", "png"), conflict_policy="auto_number",
    )

    published = atomic_write_set(
        reservation,
        {
            "csv": lambda path: path.write_text("new", encoding="utf-8"),
            "png": lambda path: path.write_bytes(b"png"),
        },
    )

    assert published["csv"].name == "result__2.csv"
    assert published["png"].name == "result__2.png"
    assert not list(tmp_path.glob(".*.batch-reserve"))


def test_atomic_write_never_replaces_a_target_that_appears_during_write(tmp_path):
    target = tmp_path / "result.csv"

    def race(temp):
        temp.write_text("ours", encoding="utf-8")
        target.write_text("racer", encoding="utf-8")

    with pytest.raises(OutputPublishRace):
        atomic_write(target, race)

    assert target.read_text(encoding="utf-8") == "racer"


def test_atomic_set_rollback_never_unlinks_outsider_replacement(
    tmp_path, monkeypatch,
):
    import mf4_analyzer.batch_output as batch_output

    reservation = reserve_output_paths(
        tmp_path, "result", ("csv", "png"), conflict_policy="auto_number",
    )
    original_publish = batch_output._publish_no_replace

    def publish_with_two_races(temp, target):
        if target.suffix == ".png":
            target.write_bytes(b"outsider-png")
            return original_publish(temp, target)
        original_publish(temp, target)
        target.unlink()
        target.write_text("outsider-csv", encoding="utf-8")

    monkeypatch.setattr(
        batch_output, "_publish_no_replace", publish_with_two_races,
    )

    with pytest.raises(OutputRollbackIncomplete, match="unknown owner"):
        atomic_write_set(
            reservation,
            {
                "csv": lambda path: path.write_text("ours", encoding="utf-8"),
                "png": lambda path: path.write_bytes(b"ours-png"),
            },
        )

    assert (tmp_path / "result.csv").read_text(encoding="utf-8") == "outsider-csv"
    assert (tmp_path / "result.png").read_bytes() == b"outsider-png"


def test_overwrite_rollback_never_restores_over_outsider_replacement(
    tmp_path, monkeypatch,
):
    import mf4_analyzer.batch_output as batch_output

    csv_path = tmp_path / "result.csv"
    png_path = tmp_path / "result.png"
    csv_path.write_text("old-csv", encoding="utf-8")
    png_path.write_bytes(b"old-png")
    reservation = reserve_output_paths(
        tmp_path, "result", ("csv", "png"), conflict_policy="overwrite",
    )
    original_replace = batch_output.os.replace

    def replace_with_outsider(source, target):
        source = Path(source)
        target = Path(target)
        is_publication = ".batch-stage." in source.name
        if is_publication and target == png_path:
            raise OSError("second artifact publish failed")
        result = original_replace(source, target)
        if is_publication and target == csv_path:
            target.unlink()
            target.write_text("outsider-csv", encoding="utf-8")
        return result

    monkeypatch.setattr(batch_output.os, "replace", replace_with_outsider)

    with pytest.raises(OutputRollbackIncomplete, match="unknown owner"):
        atomic_write_set(
            reservation,
            {
                "csv": lambda path: path.write_text("new-csv", encoding="utf-8"),
                "png": lambda path: path.write_bytes(b"new-png"),
            },
        )

    assert csv_path.read_text(encoding="utf-8") == "outsider-csv"
    assert png_path.read_bytes() == b"old-png"


def test_reservation_token_has_inspectable_owner_and_explicit_safe_release(tmp_path):
    first = reserve_output_paths(
        tmp_path, "result", ("csv",), conflict_policy="error",
    )
    info = inspect_output_reservation(first.token_path)

    assert info is not None
    assert info.metadata["schema_version"] == 1
    assert info.metadata["stem"] == "result"
    assert isinstance(info.metadata["pid"], int)
    assert info.metadata["created_at"].endswith("Z")
    with pytest.raises(
        FileExistsError, match="active or stale.*explicit inspect/release",
    ):
        reserve_output_paths(
            tmp_path, "result", ("csv",), conflict_policy="error",
        )

    release_output_reservation(info)
    replacement = reserve_output_paths(
        tmp_path, "result", ("csv",), conflict_policy="error",
    )
    replacement.release()
    first.release()


def test_explicit_reservation_release_refuses_changed_token(tmp_path):
    reservation = reserve_output_paths(
        tmp_path, "result", ("csv",), conflict_policy="error",
    )
    info = inspect_output_reservation(reservation.token_path)
    reservation.token_path.unlink()
    reservation.token_path.write_text("replacement owner", encoding="utf-8")

    with pytest.raises(RuntimeError, match="changed since inspection"):
        release_output_reservation(info)

    assert reservation.token_path.read_text(encoding="utf-8") == "replacement owner"
    reservation.release()


def test_reservation_inspection_refuses_non_token_path(tmp_path):
    unrelated = tmp_path / "measurement.csv"
    unrelated.write_text("keep", encoding="utf-8")

    with pytest.raises(ValueError, match="batch-reserve"):
        inspect_output_reservation(unrelated)

    assert unrelated.read_text(encoding="utf-8") == "keep"


def test_automatic_reservation_release_never_unlinks_replaced_token(tmp_path):
    reservation = reserve_output_paths(
        tmp_path, "result", ("csv",), conflict_policy="error",
    )
    reservation.token_path.unlink()
    reservation.token_path.write_text("outsider token", encoding="utf-8")

    reservation.release()

    assert reservation.token_path.read_text(encoding="utf-8") == "outsider token"
    reservation.token_path.unlink()


def test_batch_output_import_does_not_load_pyqt5():
    """``batch_output`` holds ``write_image``, whose Qt renderer import must
    stay lazy (inside the function body) so importing the module itself never
    pulls in PyQt5 -- design D3 / plan Task 3 Step 3.
    """
    repo_root = Path(__file__).resolve().parents[1]
    script = """
import json
import sys
import mf4_analyzer.batch_output
blocked = sorted(
    name for name in sys.modules
    if name == 'PyQt5' or name.startswith('PyQt5.')
)
print(json.dumps(blocked))
"""
    env = dict(os.environ)
    env["PYTHONPATH"] = str(repo_root)
    result = subprocess.run(
        [sys.executable, "-c", script],
        cwd=repo_root,
        env=env,
        text=True,
        capture_output=True,
        check=False,
    )

    assert result.returncode == 0, result.stderr
    assert json.loads(result.stdout) == []
